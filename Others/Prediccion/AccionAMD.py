import yfinance as yf
import openpyxl
import datetime

# Ruta
file_path = "./Others/Prediccion/predictions.xlsx"      
# Abrir el archivo de Excel
workbook = openpyxl.load_workbook(file_path)

# Obtener información de la acción de AMD
ticker = yf.Ticker("AMD")

# Obtener el precio de cierre de la acción de AMD en el último mes
history = ticker.history(period="1mo")
closing_prices = history["Close"]

# Calcular el promedio del movimiento del precio de cierre en el último mes
movement_average = sum(closing_prices) / len(closing_prices)

# Obtener la fecha de hoy y mañana en formato "dd-mm-yyyy"
tomorrow = (datetime.datetime.now() + datetime.timedelta(days=1)).strftime("%Y-%m-%d")
today = datetime.datetime.today().strftime("%Y-%m-%d")

# Seleccionar la primera hoja del libro
worksheet = workbook.active
# Obtener el número de filas en la hoja
row_count = worksheet.max_row

# Recorrer todas las filas de la hoja
for row in worksheet.iter_rows(min_row=2): 
    # Recorrer los índices de tiempo del DataFrame   
    for index, check in history.iterrows():
        # Verificar si el índice de tiempo (fecha) coincide con la fecha que se quiere buscar
        if index.strftime("%Y-%m-%d") == row[1].value:
            # Si la celda de la columna "Real" está vacía y es diferente al dia actual
            if row[3].value is None and row[1].value != tomorrow:
                row[3].value = check["Close"]

# Añadir una nueva fila al final de la hoja
worksheet.append([row_count+1, tomorrow, movement_average])
# Guardar los cambios en el archivo
workbook.save(file_path)

print("Predicción del precio de cierre del próximo día guardada en el archivo de Excel")